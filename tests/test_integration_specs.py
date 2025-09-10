# tests/test_integration_specs.py
from pathlib import Path
import importlib
import sys

import openpyxl
import pytest


# ---- helpers ---------------------------------------------------------------

def run_cli(monkeypatch, argv):
    """Import cli module with patched argv each run."""
    monkeypatch.setenv("PYTHONIOENCODING", "utf-8")
    monkeypatch.setenv("PYTHONUNBUFFERED", "1")
    monkeypatch.setattr(sys, "argv", argv)
    cli = importlib.import_module("api_description_tool.cli")
    importlib.reload(cli)
    return cli


def _write_cfg_xlsx(tmp_path, validate=False):
    cfg = tmp_path / "config.ini"
    cfg.write_text(
        f"""
[input]
validate={str(validate)}

[output]
format=xlsx
""".strip(),
        encoding="utf-8",
    )
    return cfg


def _data_file(name: str) -> Path:
    # repo/tests/data/<name>
    here = Path(__file__).parent
    p = here / "data" / name
    if not p.exists():
        pytest.skip(f"Missing test data file: {p} — place your YAMLs under tests/data/")
    return p


def _assert_excel_ok(xlsx: Path):
    assert xlsx.exists(), f"Excel file missing: {xlsx}"
    wb = openpyxl.load_workbook(str(xlsx))
    assert wb.sheetnames == ["Params", "Req Body", "Res Body"]

    # header checks and at least one data row (>=2 rows total)
    params_h = [c.value for c in wb["Params"][1]]
    req_h = [c.value for c in wb["Req Body"][1]]
    res_h = [c.value for c in wb["Res Body"][1]]

    assert params_h[:6] == ["Name", "Mandatory", "Expected Value(s)", "In", "Description", "Examples"]
    assert req_h[:6] == ["Path", "Property", "Mandatory", "Expected Value(s)", "Description", "Examples"]
    assert res_h[:7] == ["Status", "Path", "Property", "Mandatory", "Expected Value(s)", "Description", "Examples"]

    assert wb["Params"].max_row >= 2
    assert wb["Req Body"].max_row >= 2
    assert wb["Res Body"].max_row >= 2


# ---- tests ----------------------------------------------------------------

def test_integration_alt_trans_an_to_xlsx(tmp_path, monkeypatch):
    """
    Full E2E: derive output base from input filename, build Excel, check sheets & headers.
    Uses tests/data/alt-trans-an.yml
    """
    spec = _data_file("alt-trans-an.yml")
    cfg = _write_cfg_xlsx(tmp_path, validate=False)

    # write into a clean temp working directory
    monkeypatch.chdir(tmp_path)
    cli = run_cli(monkeypatch, ["prog", str(spec), "--config", str(cfg)])
    cli.main()

    expected = tmp_path / f"{spec.stem}_api_tab_desc.xlsx"
    _assert_excel_ok(expected)


def test_integration_psc_an_to_xlsx(tmp_path, monkeypatch):
    """
    Full E2E: derive output base from input filename, build Excel, check sheets & headers.
    Uses tests/data/psc_an.yml
    """
    spec = _data_file("psc_an.yml")
    cfg = _write_cfg_xlsx(tmp_path, validate=False)

    monkeypatch.chdir(tmp_path)
    cli = run_cli(monkeypatch, ["prog", str(spec), "--config", str(cfg)])
    cli.main()

    expected = tmp_path / f"{spec.stem}_api_tab_desc.xlsx"
    _assert_excel_ok(expected)
