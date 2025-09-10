from __future__ import annotations

from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Font


PARAMS_HEADERS = ["Name", "Mandatory", "Expected Value(s)", "In", "Description", "Examples"]
REQ_HEADERS = ["Path", "Property", "Mandatory", "Expected Value(s)", "Description", "Examples"]
RES_HEADERS = ["Status", "Path", "Property", "Mandatory", "Expected Value(s)", "Description", "Examples"]


def _truthy(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().lower() in {"1", "true", "yes", "on"}


def _write_sheet(ws, headers: List[str], rows: List[Dict[str, object]], *, bold_fields: List[str]):
    # headers
    ws.append(headers)
    # rows
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    # Identify columns
    col_index = {h: i + 1 for i, h in enumerate(headers)}  # 1-based
    mandatory_idx = col_index.get("Mandatory")

    # Bold requested fields when Mandatory == True
    if mandatory_idx:
        for r in range(2, ws.max_row + 1):  # skip header
            is_mand = _truthy(ws.cell(row=r, column=mandatory_idx).value)
            if is_mand:
                for field in bold_fields:
                    cidx = col_index.get(field)
                    if cidx:
                        cell = ws.cell(row=r, column=cidx)
                        cell.font = Font(bold=True)

    # Auto width (simple)
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)


def write_excel(file_path: str, params_rows: List[Dict[str, object]], req_rows: List[Dict[str, object]], res_rows: List[Dict[str, object]]):
    wb = Workbook()

    # Params: bold Name for mandatory (no Path column here)
    ws_params = wb.active
    ws_params.title = "Params"
    _write_sheet(ws_params, PARAMS_HEADERS, params_rows, bold_fields=["Name"])

    # Req Body: bold both Path and Property on mandatory rows
    ws_req = wb.create_sheet("Req Body")
    _write_sheet(ws_req, REQ_HEADERS, req_rows, bold_fields=["Path", "Property"])

    # Res Body: bold both Path and Property on mandatory rows
    ws_res = wb.create_sheet("Res Body")
    _write_sheet(ws_res, RES_HEADERS, res_rows, bold_fields=["Path", "Property"])

    wb.save(file_path)